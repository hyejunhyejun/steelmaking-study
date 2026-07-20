import os, io, json, datetime
import openpyxl
from PIL import Image
from parse import parse_rows, MAIN_RE, SUB_RE
from derive import derive_keywords, compute_groups, normalize

MAX_IMG_WIDTH = 1000  # 큰 스캔 이미지를 모바일 친화적 크기로 축소

# 엑셀에 회전된 채 삽입된 그림을 바로 세우는 맵. 키=f"{sheet}_{index}", 값=CCW 각도.
# (전수 확인 결과: 부두아 곡선·슬래그 삼원계·소결 공정도는 +90, 풍구 단면은 -90)
ROTATE = {
    "21-1_1": 90, "21-2_1": 90, "22-1_1": 90, "23-1_0": 90,
    "24-1_1": 90, "24-2_0": 90, "25-1_2": 90, "25-2_0": 90,  # 부두아(8장 동일)
    "21-1_3": 90, "24-2_1": 90,                               # 슬래그 삼원계
    "21-1_4": 90,                                             # 소결 공정도
    "21-2_0": -90,                                            # 풍구 단면
}

ROUND_SHEETS = ["21-1", "21-2", "22-1", "22-2", "23-1", "23-2",
                "24-1", "24-2", "25-1", "25-2"]
WRONG_SHEET = "틀린문제"


def _rows(ws):
    out = []
    for r in range(1, ws.max_row + 1):
        out.append((ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value))
    return out


def _question_start_rows(ws):
    """대문제가 시작하는 행 번호(파서의 MAIN 판정과 1:1 일치)."""
    idx = []
    for r in range(1, ws.max_row + 1):
        c = ws.cell(r, 3).value
        cs = c.strip() if isinstance(c, str) else None
        if cs and MAIN_RE.match(cs) and not SUB_RE.match(cs):
            idx.append(r)
    return idx


def _extract_images(ws, sheet_id, out_dir, starts):
    """이미지 저장 후 {대문제 시작행: [상대경로,...]} 반환."""
    os.makedirs(out_dir, exist_ok=True)
    mapping = {}
    for i, img in enumerate(getattr(ws, "_images", [])):
        try:
            anchor_row = img.anchor._from.row + 1
        except Exception:
            continue
        owner = max([s for s in starts if s <= anchor_row], default=None)
        key = f"{sheet_id}_{i}"
        fname = key + ".jpg"
        _save_optimized(img._data(), os.path.join(out_dir, fname), ROTATE.get(key, 0))
        mapping.setdefault(owner, []).append(f"images/{fname}")
    return mapping


def _save_optimized(raw_bytes, path, rotate=0):
    """스캔 이미지를 (필요 시 회전 후) 최대 폭 이하로 축소하고 JPEG로 재저장."""
    try:
        im = Image.open(io.BytesIO(raw_bytes)).convert("RGB")
        if rotate:
            im = im.rotate(rotate, expand=True)
        if im.width > MAX_IMG_WIDTH:
            h = round(im.height * MAX_IMG_WIDTH / im.width)
            im = im.resize((MAX_IMG_WIDTH, h), Image.LANCZOS)
        im.save(path, "JPEG", quality=82, optimize=True)
    except Exception:
        with open(os.path.splitext(path)[0] + ".png", "wb") as f:
            f.write(raw_bytes)


def build_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    images_dir = os.path.join(os.path.dirname(path), "..", "data", "images")
    rounds = []
    exam_items = []   # (round_id, text) — 별표는 회차(exam)만으로 계산
    all_q_refs = []   # 별표/그룹을 다시 꽂을 전체 문제 참조

    def build_round(sheet_name, rid, label, rtype):
        ws = wb[sheet_name]
        questions = parse_rows(_rows(ws))
        starts = _question_start_rows(ws)
        img_map = _extract_images(ws, rid, images_dir, starts)
        for q, start_row in zip(questions, starts):
            q["images"] = img_map.get(start_row, [])
            for part in q["parts"]:
                part["keywords"] = derive_keywords(part["answers"])
            if rtype == "exam":
                exam_items.append((rid, q["text"]))
            all_q_refs.append(q)
        return {"id": rid, "label": label, "type": rtype, "questions": questions}

    for s in ROUND_SHEETS:
        yy, n = s.split("-")
        rounds.append(build_round(s, s, f"20{yy}년 {n}회", "exam"))
    rounds.append(build_round(WRONG_SHEET, "wrong", "오답노트", "wrong"))

    # 회차 문제로만 별표 그룹 계산 → 정규화 텍스트 기준 사전
    groups = compute_groups(exam_items)
    by_text = {}
    for (rid, text), g in zip(exam_items, groups):
        by_text[normalize(text)] = g

    for q in all_q_refs:
        key = normalize(q["text"])
        g = by_text.get(key)
        if g:
            q["stars"] = g["stars"]
            q["groupId"] = g["groupId"]
        else:
            q["stars"] = 1
            q["groupId"] = "g-" + key[:24]

    return {"meta": {"generatedAt": datetime.date.today().isoformat()}, "rounds": rounds}


def main():
    here = os.path.dirname(__file__)
    xlsx = os.path.join(here, "..", "원본자료", "제선기능장.xlsx")
    data = build_workbook(xlsx)
    out = os.path.join(here, "..", "data", "questions.json")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    n = sum(len(r["questions"]) for r in data["rounds"])
    print(f"작성 완료: {out} (라운드 {len(data['rounds'])}개, 문제 {n}개)")


if __name__ == "__main__":
    main()
