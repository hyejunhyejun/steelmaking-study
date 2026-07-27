import os, io
import openpyxl
import numpy as np
from PIL import Image, ImageFilter
# 문제 번호 판정은 워드 파서와 반드시 같은 규칙을 써야 한다.
# (회차, 번호) 조인이 어긋나면 그림이 엉뚱한 문제에 붙는다.
from parse_docx import MAIN_RE, SUB_RE

MAX_IMG_WIDTH = 1000
# 원본이 너무 작은 스캔은 화면에서 읽기 어려워 적당히 키운다(과도한 확대는 금지)
MIN_DISPLAY_WIDTH = 460
MIN_DISPLAY_HEIGHT = 260
MAX_UPSCALE = 2.2
ROUND_SHEETS = ["21-1", "21-2", "22-1", "22-2", "23-1", "23-2",
                "24-1", "24-2", "25-1", "25-2"]

# 엑셀에 회전된 채 삽입된 그림을 바로 세우는 맵(전수 육안 확인 완료). 값=반시계 각도.
ROTATE = {
    "21-1_1": 90, "21-2_1": 90, "22-1_1": 90, "23-1_0": 90,
    "24-1_1": 90, "24-2_0": 90, "25-1_2": 90, "25-2_0": 90,  # 부두아(8장 동일)
    "21-1_3": 90, "24-2_1": 90,                               # 슬래그 삼원계
    "21-1_4": 90,                                             # 소결 공정도
    "21-2_0": 90,                                             # 풍구 단면
    "24-2_3": 90, "25-2_1": 90, "25-2_2": 90,                 # 폰캡처 내부 눕힘
    "24-1_0": 90,                                             # 열풍로 배관도(눕힘)
    "24-2_2": 90,                                             # 노정장입물 4패널(눕힘)
}
# 휴대폰 촬영 앱 화면째로 캡처된 이미지 — 상단/하단 UI 바 제거(iOS 고정 위치)
CROP_PHONE_UI = {"22-1_0", "24-1_0", "24-2_2", "24-2_3", "25-2_1", "25-2_2"}
CROP_TOP, CROP_BOTTOM = 130, 115
# UI를 잘라낸 뒤에도 남는 회색 배경·빈 여백을 잉크 영역까지 다듬는다
INK_THRESHOLD = 170   # 이 값보다 어두우면 잉크(선·글자)로 본다
INK_MIN_PIXELS = 3    # 노이즈 무시용 최소 잉크 픽셀 수
TRIM_PADDING = 10

# 스캔·촬영으로 기울어진 그림을 자동으로 바로 세운다.
# 행별 잉크량의 분산이 최대가 되는 각도를 찾는다(수평선이 정렬될 때 최대).
DESKEW_RANGE = [i * 0.5 for i in range(-10, 11)]   # -5.0 ~ +5.0도
DESKEW_MIN_GAIN = 0.06   # 이 비율 이상 개선될 때만 적용(오작동 방지)

# 그림에 문제 지문이 함께 촬영된 경우 — 지문을 잘라내고 그림만 남긴다.
# 값은 (좌, 상, 우, 하) 비율. 트림 후 최종 단계에 적용.
TEXT_CROP = {
    "24-2_3": (0.00, 0.40, 1.00, 0.74),   # '20 다음의 안전 표지판에…' 지문 제거
    "25-2_2": (0.20, 0.20, 1.00, 1.00),   # '19 다음 그림은 소결의 급광장치…' 지문 제거
}


def _question_rows(ws):
    """(행번호, 대문제번호) 목록."""
    out = []
    for r in range(1, ws.max_row + 1):
        c = ws.cell(r, 3).value
        cs = c.strip() if isinstance(c, str) else None
        if cs and MAIN_RE.match(cs) and not SUB_RE.match(cs):
            out.append((r, int(MAIN_RE.match(cs).group(1))))
    return out


def _ink_bbox(im):
    """잉크(어두운 픽셀)가 차지하는 영역 + 여백. 없으면 None."""
    g = np.asarray(im.convert("L"))
    ink = g < INK_THRESHOLD
    rows = np.where(ink.sum(axis=1) >= INK_MIN_PIXELS)[0]
    cols = np.where(ink.sum(axis=0) >= INK_MIN_PIXELS)[0]
    if len(rows) == 0 or len(cols) == 0:
        return None
    h, w = g.shape
    p = TRIM_PADDING
    return (max(0, int(cols.min()) - p), max(0, int(rows.min()) - p),
            min(w, int(cols.max()) + p + 1), min(h, int(rows.max()) + p + 1))


def _strip_dark_borders(im, dark=60, ratio=0.9):
    """스크린샷 레터박스(균일한 검은 띠)를 잘라낸다."""
    g = np.asarray(im.convert("L"))
    h, w = g.shape
    top, bottom, left, right = 0, h, 0, w
    while top < bottom and (g[top, left:right] < dark).mean() > ratio:
        top += 1
    while bottom > top and (g[bottom - 1, left:right] < dark).mean() > ratio:
        bottom -= 1
    while left < right and (g[top:bottom, left] < dark).mean() > ratio:
        left += 1
    while right > left and (g[top:bottom, right - 1] < dark).mean() > ratio:
        right -= 1
    if (left, top, right, bottom) == (0, 0, w, h):
        return im
    return im.crop((left, top, right, bottom))


def _row_ink_variance(im):
    """행별 잉크량의 분산 — 수평선이 정렬될수록 커진다."""
    g = np.asarray(im.convert("L"))
    return float((g < INK_THRESHOLD).sum(axis=1).astype(float).var())


def _deskew(im):
    """기울어진 스캔을 바로 세운다(개선폭이 작으면 원본 유지)."""
    base = _row_ink_variance(im)
    if base <= 0:
        return im
    best_angle, best_score = 0.0, base
    for a in DESKEW_RANGE:
        if a == 0:
            continue
        cand = im.rotate(a, expand=False, fillcolor=(255, 255, 255))
        s = _row_ink_variance(cand)
        if s > best_score:
            best_angle, best_score = a, s
    if best_angle == 0 or best_score / base - 1 < DESKEW_MIN_GAIN:
        return im
    # 잘림 없이 회전한 뒤 남는 여백은 뒤이은 트림에서 제거된다
    return im.rotate(best_angle, expand=True, fillcolor=(255, 255, 255))


def _upscale_small(im):
    """너무 작은 스캔을 읽기 좋은 크기로 키우고 선명화한다."""
    f = max(MIN_DISPLAY_WIDTH / im.width, MIN_DISPLAY_HEIGHT / im.height)
    f = min(f, MAX_UPSCALE)
    if f <= 1.01:
        return im
    im = im.resize((round(im.width * f), round(im.height * f)), Image.LANCZOS)
    return im.filter(ImageFilter.UnsharpMask(radius=1.2, percent=110, threshold=3))


def _straighten(im):
    """여백 트림 → 기울기 보정 → 재트림. 여백이 크면 기울기 측정이 흐려진다."""
    for _ in range(2):
        bb = _ink_bbox(im)
        if bb:
            im = im.crop(bb)
        im = _deskew(im)
    bb = _ink_bbox(im)
    return im.crop(bb) if bb else im


def _save(raw_bytes, path, rotate=0, crop_ui=False, text_crop=None):
    """폰 UI 크롭 → 90도 회전 → 기울기 보정 → 잉크 트림 → 지문 제거 → 축소 → 저장."""
    try:
        im = Image.open(io.BytesIO(raw_bytes)).convert("RGB")
        if crop_ui:
            w, h = im.size
            im = im.crop((0, CROP_TOP, w, h - CROP_BOTTOM))
        if rotate:
            im = im.rotate(rotate, expand=True)
        im = _strip_dark_borders(im)
        im = _straighten(im)
        if text_crop:
            w, h = im.size
            l, t, r, b = text_crop
            im = im.crop((int(w * l), int(h * t), int(w * r), int(h * b)))
            # 지문을 떼어낸 뒤 다시 재보정(남은 그림 기준으로 기울기가 달라진다)
            im = _straighten(im)
        im = _upscale_small(im)
        if im.width > MAX_IMG_WIDTH:
            h2 = round(im.height * MAX_IMG_WIDTH / im.width)
            im = im.resize((MAX_IMG_WIDTH, h2), Image.LANCZOS)
        im.save(path, "JPEG", quality=82, optimize=True)
    except Exception:
        with open(os.path.splitext(path)[0] + ".png", "wb") as f:
            f.write(raw_bytes)


def extract_images(xlsx_path, out_dir):
    """엑셀 그림을 저장하고 {(회차id, 대문제번호): [경로,...]} 반환."""
    os.makedirs(out_dir, exist_ok=True)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    mapping = {}
    for sheet in ROUND_SHEETS:
        ws = wb[sheet]
        rows = _question_rows(ws)
        for i, img in enumerate(getattr(ws, "_images", [])):
            try:
                anchor = img.anchor._from.row + 1
            except Exception:
                continue
            owner = None
            for r, num in rows:
                if r <= anchor:
                    owner = num
            if owner is None:
                continue
            key = f"{sheet}_{i}"
            _save(img._data(), os.path.join(out_dir, key + ".jpg"),
                  ROTATE.get(key, 0), key in CROP_PHONE_UI, TEXT_CROP.get(key))
            mapping.setdefault((sheet, owner), []).append(f"images/{key}.jpg")
    return mapping


# ---------------------------------------------------------------------------
# 워드(사진정리본)에 내장된 실제 사진 추출
# ---------------------------------------------------------------------------
_A_NS = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
_R_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
_Q_RE = None  # 지연 컴파일(모듈 상단 import 순서 유지)


def extract_docx_photos(docx_path, out_dir, wanted_nums):
    """사진정리본 워드에서 지정한 '문제 N'의 내장 사진을 저장.

    반환: {문제번호: "images/파일명"} — 없는 번호는 결과에 없다.
    """
    import re
    import docx as _docx

    global _Q_RE
    if _Q_RE is None:
        _Q_RE = re.compile(r"^문제\s*(\d+)\.")

    os.makedirs(out_dir, exist_ok=True)
    doc = _docx.Document(docx_path)
    cur, found = None, {}
    for p in doc.paragraphs:
        m = _Q_RE.match(p.text.strip())
        if m:
            cur = int(m.group(1))
        if cur not in wanted_nums:
            continue
        for blip in p._p.findall(f".//{_A_NS}blip"):
            rid = blip.get(f"{_R_NS}embed")
            if not rid:
                continue
            blob = doc.part.related_parts[rid].blob
            name = f"photo_{cur}.jpg"
            _save(blob, os.path.join(out_dir, name))
            found[cur] = f"images/{name}"
    return found


# ---------------------------------------------------------------------------
# 추가 노트 사진(29장)에서 특정 그림만 잘라 쓰기
# ---------------------------------------------------------------------------
# 화면을 찍은 사진이라 회색빛·모아레가 있어 색을 유지한 채 대비를 올린다.
NOTE_FIGURES = {
    # 저장이름: (사진 인덱스(0부터), 자를 영역(좌,상,우,하))
    "note_quartz": (4, (100, 128, 562, 362)),    # 5쪽 석영 상변태 곡선
}


def _enhance_screen_photo(im):
    from PIL import ImageOps, ImageEnhance
    im = ImageOps.autocontrast(im, cutoff=(1, 10))
    im = ImageEnhance.Color(im).enhance(1.3)
    return ImageEnhance.Contrast(im).enhance(1.15)


def extract_note_figures(notes_dir, out_dir):
    """노트 사진에서 지정한 그림을 잘라 저장하고 {이름: 상대경로} 반환."""
    if not os.path.isdir(notes_dir):
        return {}
    files = sorted(os.listdir(notes_dir))
    os.makedirs(out_dir, exist_ok=True)
    made = {}
    for name, (idx, box) in NOTE_FIGURES.items():
        if idx >= len(files):
            continue
        im = Image.open(os.path.join(notes_dir, files[idx])).convert("RGB").crop(box)
        buf = io.BytesIO()
        _enhance_screen_photo(im).save(buf, "PNG")
        _save(buf.getvalue(), os.path.join(out_dir, name + ".jpg"))
        made[name] = f"images/{name}.jpg"
    return made
